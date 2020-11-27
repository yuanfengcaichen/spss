import copy
import threading
import uuid
import time
import statsmodels.api as sm
import pandas as pd
import numpy as np
import requests
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from rest_framework.views import APIView
import json
from django.core.cache import cache #引入redis缓存
from django_redis import get_redis_connection
import base64
import xlrd
from analysis.tools.mydecode import mapkeydecode
from analysis.tools.myredis import getconn
import pickle
from analysis.tools.mytranslate import tranmodel
import matplotlib.pyplot as plt
from pylab import *
import base64
from io import BytesIO
# Create your views here.
'''
from analysis.linear.curmodel import setcurmodel
from analysis.linear.linearcorrelationgraph import linear_correlation
from analysis.linear.model import setmodel
from analysis.linear.outliertest import outliertest
from analysis.linear.ppqqgraph import pp, qq
from analysis.linear.prediction import prediction
from analysis.linear.regression import analysis, returncloumns, normality, multicol, norks

from openpyxl import load_workbook

#fileList=[]
from analysis.linear.residual import residual
from analysis.linear.variance import variance, varbp
from analysis.models import Red
'''



def getaddress(id,ip):
    url = 'https://api.map.baidu.com/location/ip?ak=rGa0BEvgESYRDkgTLSIwkwHN5zkLfGcA&ip='+ip+'&coor=bd09ll'  # 请求接口
    req = requests.get(url)#发送请求
    data = req.json()
    #print(ip + "----" + data.get("content").get("address"))#获取请求，得到的是字典格式

def savefile(file,sheet,fid):
    from analysis.linear.regression import returncloumns
    p = returncloumns(file, sheet)

    #使用redis进行存、取数据
    conn = getconn()
    #存数据
    data={}
    data["Profit"] = pickle.dumps(p)
    conn.hmset(fid, data)
    conn.expire(fid, 60*60*2)
    #取数据
    # map = getconn().hgetall(fid)#二进制数据
    # newmap = mapkeydecode(map)
    # data = pickle.loads(newmap.get('Profit'))
    # print(type(data))

Files={}
class index(APIView):
    def get(self, request, *args, **kwargs):
        return render(request, "index.html")

# def index(request):#返回多元线性回归网页
#     # x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')  # 判断是否使用代理
#     # if x_forwarded_for:
#     #     ip = x_forwarded_for.split(',')[0]  # 使用代理获取真实的ip
#     # else:
#     #     ip = request.META.get('REMOTE_ADDR')  # 未使用代理获取IP
#     # t1 = threading.Thread(target=getaddress, args=(1,ip))  # 新开一个线程获取访问地址
#     # t1.start()
#     # l = pd.DataFrame(list(Red.objects.all())) # 使用数据库读取数据
#     # print(type(l))
#     # for a in l:
#     #     pass
#     return render(request, 'index.html')

class linear(APIView):
    def get(self, request, *args, **kwargs):
        return render(request, "regression_index.html")
class gradually(APIView):
    def get(self, request, *args, **kwargs):
        return render(request, "regression_index.html")

def uploadfile(request):#用户上传文件，返回文件中的列名
    from analysis.linear.regression import returncloumns
    file = request.FILES.get("file")
    filename = file.name
    time_start = time.time()
    #print(time_start)
    sheets = pd.ExcelFile(file).sheet_names
    #print(sheets)
    time_end = time.time()
    #print('totally cost: ', time_end - time_start)
    resultdatas=[]
    for sheet in sheets:
        uid = str(uuid.uuid4())
        fid = ''.join(uid.split('-'))
        t1 = threading.Thread(target=savefile, args=(file, sheet, fid))  # 新开一个线程保存读取的文件
        t1.start()
        p = returncloumns(file,sheet)
        columns = p.columns.values.tolist()
        resultdata = [filename + '-' + sheet, columns, fid]
        resultdatas.append(resultdata)
        time_end = time.time()
        #print('totally cost: ', time_end - time_start)
    #print(resultdatas)
    ret1 = json.loads(json.dumps(resultdatas, ensure_ascii=False))
    return JsonResponse({"result": 1,"resultdata":ret1}, json_dumps_params={'ensure_ascii': False})

"""
def uploadfile(request):#用户上传文件，返回文件中的列名
    if request.method == "POST":
        from openpyxl import load_workbook
        global Files
        file = request.FILES.get("file")
        filename = file.name
        tables = load_workbook(file)
        sheets = tables.sheetnames
        resultdatas=[]
        for sheet in sheets:
            uid = str(uuid.uuid4())
            fid = ''.join(uid.split('-'))
            copyfile = copy.deepcopy(file)
            t1 = threading.Thread(target=savefile, args=(copyfile,sheet,fid))  # 新开一个线程保存读取的文件
            t1.start()
            table = tables.get_sheet_by_name(sheet)
            a = table.max_column
            columns = []
            for i in range(1, a + 1):
                columns.append(table.cell(row=1,column=i).value)
            resultdata=[filename+'-'+sheet,columns,fid]
            resultdatas.append(resultdata)
        ret1 = json.loads(json.dumps(resultdatas, ensure_ascii=False))
        return JsonResponse({"result": 1,"resultdata":ret1}, json_dumps_params={'ensure_ascii': False})
"""

def sendselect(request):#用户选择x轴和y轴，进行回归分析，返回模型数据
    if request.method == "POST":
        #print(request.body)
        data=json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        analytype = data["analytype"]
        criterion = data["criterion"]
        direction = data["direction"]
        conn = getconn()
        if (conn.exists(fileindex)):
            conn.expire(fileindex, 60*60*2)
            if(analytype=="linear" and conn.hexists(fileindex,'xselected_change')):
                conn.hdel(fileindex,'xselected_change')
            return sendselecthelp(fileindex, xselected, yselected, analytype, criterion, direction)
        else:
            responsedata = {"result": 404,"msg":'上传的文件已过期，请重新上传'}
            return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def sendselecthelp(fileindex, xselected, yselected, analytype, criterion, direction):
    from analysis.linear.curmodel import setcurmodel
    from analysis.linear.model import setmodel
    from analysis.linear.regression import analysis
    t1 = threading.Thread(target=setmodel, args=(
        fileindex, xselected, yselected, analytype, criterion, direction))  # 新开一个线程获取模型
    t1.start()
    t1.join()
    t2 = threading.Thread(target=setcurmodel, args=(fileindex, xselected, yselected))  # 新开一个线程获取修正后的模型
    t2.start()
    f = analysis(fileindex, xselected, yselected, analytype, criterion, direction)
    model_summary = tranmodel(f.get('model').summary().as_html())
    model = json.loads(json.dumps(model_summary, ensure_ascii=False))
    model_params = json.loads(json.dumps(f.get('model').params.index.tolist(), ensure_ascii=False))
    responsedata = {"result": 1, "model": model, "model_params":model_params, "f1": f.get('f1'), "f2": f.get('f2'),'xselected_change':f.get('xselected_change')}
    return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getsin_pre_value(request):#获取模型预测值
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        params = list(map(float,data["params"]))
        conn = getconn()
        est = pickle.loads(conn.hget(fileindex, 'est'))
        if 'const' in est.params.index.tolist():
            params.insert(0, 1)
        params = np.array(params)
        sin_pre_value = est.params.values * params
        sin_pre_value = sin_pre_value.sum()
        sin_pre_value = {"result": 1, "sin_pre_value": round(sin_pre_value, 3)}
        return JsonResponse(sin_pre_value, json_dumps_params={'ensure_ascii': False})

def uploadpre_file(request):#上传多值预测文件
    import matplotlib
    matplotlib.use('Agg')
    from matplotlib import pyplot as plt
    file = request.FILES.get("file")
    filename = file.name
    fileindex = request.POST.get("fileindex")
    xselected = request.POST.get("xselected")
    yselected = request.POST.get("yselected")
    conn = getconn()
    Profit = pd.read_excel(file)
    Profit.dropna(inplace=True)
    est = pickle.loads(conn.hget(fileindex, 'est'))
    params = est.params.index.tolist()
    if 'const' in params:
        params.remove('const')
    try:
        Profit = Profit[params]
        Profit = sm.add_constant(Profit)
        y_pred = est.predict(Profit)
        Profit[yselected + "(预测值)"] = y_pred
        if 'const' in Profit.columns.values.tolist():
            Profit = Profit.drop('const', axis=1)
        Profit = round(Profit, 3)
        Profit = Profit.to_dict('records')
        # 散点图
        plt.scatter(range(1, len(y_pred) + 1), y_pred, alpha=0.4, edgecolor='none')
        sio = BytesIO()
        plt.savefig(sio, format='png', bbox_inches='tight', pad_inches=0.0)
        data = base64.encodebytes(sio.getvalue()).decode()
        src = 'data:image/png;base64,' + str(data)
        # 记得关闭，不然画出来的图是重复的
        plt.axis('off')
        plt.close()
        #Profit = [{"a":1,"b":2,"c":3},{"a":4,"b":5,"c":6}]
        mul_pre_result = {"mul_pre_values": Profit, "src": src}
        #ret1 = json.loads(json.dumps(Profit, ensure_ascii=False))
        return JsonResponse({"result": 1, "mul_pre_result": mul_pre_result}, json_dumps_params={'ensure_ascii': False})
    except KeyError as e:
        data = {"result": 500, "except": "keyerror"}
        return JsonResponse(data, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        data = {"result": 500, "except": "error"}
        return JsonResponse(data, json_dumps_params={'ensure_ascii': False})
    except BaseException as e:
        print(e)



def getprediction(request):#获取模型预测图片
    from analysis.linear.prediction import prediction
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        prediction_src = prediction(fileindex,xselected,yselected)
        responsedata = {"prediction_src": prediction_src}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getnormality(request):#获取正态性检验的图片
    from analysis.linear.regression import normality
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        normality_src = normality(fileindex, yselected)
        responsedata = {"normality_src":normality_src}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getppqq(request):#获取qqpp图片地址
    from analysis.linear.ppqqgraph import pp, qq
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        pp_src = pp(fileindex, yselected)
        qq_src = qq(fileindex, yselected)
        responsedata = {"pp_src":pp_src,"qq_src":qq_src,}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getks(request):
    from analysis.linear.regression import norks
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        ks = norks(fileindex,yselected)
        responsedata = {"ks": ks}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getmulticol(request):#获取多重共线性表格数据
    from analysis.linear.regression import multicol
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        multicollinearity = multicol(fileindex, xselected)
        multicollinearity = json.loads(json.dumps(multicollinearity, ensure_ascii=False))
        responsedata = {"multicollinearity":multicollinearity}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getlinearcorrelate(request):#获取线性相关性图片
    from analysis.linear.linearcorrelationgraph import linear_correlation
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        lineselected = data["lineselected"]
        linear_correlation_src = linear_correlation(fileindex,lineselected)
        responsedata = {"linear_correlation_src": linear_correlation_src}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getoutliertest(request):#获取异常值检测模型
    from analysis.linear.outliertest import outliertest
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        f = outliertest(fileindex, xselected, yselected)
        model_summary = tranmodel(f.get('model'))
        testmodel = json.loads(json.dumps({'model': model_summary, 'outdata': f.get('outdata'), 'src': f.get('src')}, ensure_ascii=False))
        responsedata = {"result": 1, "testmodel": testmodel, }
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getresidual(request):#获取残差独立性相关数据
    from analysis.linear.residual import residual
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        dw = residual(fileindex, xselected, yselected)
        responsedata = {"dw": dw}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getbp(request):
    from analysis.linear.variance import varbp
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        bp = varbp(fileindex, xselected, yselected)
        responsedata = {"bp": bp}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})

def getvariance(request):#获取方差齐性检验图片：
    from analysis.linear.variance import variance
    if request.method == "POST":
        data = json.loads(request.body)
        fileindex = data["fileindex"]
        xselected = data["xselected"]
        yselected = data["yselected"]
        oselected_1 = data["oselected_1"]
        oselected_2 = data["oselected_2"]
        variance_src = variance(fileindex, xselected, yselected,oselected_1,oselected_2)
        responsedata = {"variance_src": variance_src}
        return JsonResponse(responsedata, json_dumps_params={'ensure_ascii': False})


